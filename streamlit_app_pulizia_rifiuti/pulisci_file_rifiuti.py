from __future__ import annotations

import argparse
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


REQUIRED_COLUMNS = ("Nr. Doc.", "MC", "Kg")

NR_DOC_NUMBER_FORMAT = "#,##0"
DECIMAL_NUMBER_FORMAT = "#,##0.00"


@dataclass(frozen=True)
class HeaderDetectionResult:
    columns: dict[str, int]
    header_row: int


@dataclass(frozen=True)
class DocCorrection:
    first_excel_row: int
    last_excel_row: int
    original_value: int
    corrected_value: int
    previous_distinct_value: int | None
    next_distinct_value: int | None
    rule: str
    affected_rows: int


@dataclass(frozen=True)
class ValueCorrection:
    excel_row: int
    column_name: str
    original_value: Any
    corrected_value: str


@dataclass(frozen=True)
class CleaningReport:
    sheet_name: str
    header_row: int
    data_rows: int
    nr_doc_corrections: list[DocCorrection]
    mc_corrections: list[ValueCorrection]
    kg_corrections: list[ValueCorrection]


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def normalize_header(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def simplify_header(value: Any) -> tuple[str, list[str]]:
    """
    Normalizza un'intestazione per riconoscere varianti come:
    - Nr. Doc., nr. doc., n. doc., Nr doc
    - MC, mc, metri cubi
    - Kg, kg, KG
    """
    text = normalize_header(value)
    text = text.replace("³", "3").replace("²", "2")

    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()

    token_text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    tokens = token_text.split() if token_text else []
    compact = "".join(tokens)

    return compact, tokens


def identify_required_column(header_value: Any) -> str | None:
    compact, tokens = simplify_header(header_value)

    if not compact:
        return None

    doc_exact = {
        "nrdoc",
        "ndoc",
        "nodoc",
        "nrodoc",
        "numdoc",
        "numerodoc",
        "docnr",
        "docn",
        "docno",
        "docnum",
        "docnumero",
        "nrdocumento",
        "ndocumento",
        "nodocumento",
        "nrodocumento",
        "numdocumento",
        "numerodocumento",
        "documentonr",
        "documenton",
        "documentono",
        "documentonum",
        "documentonumero",
    }

    doc_tokens = {"doc", "documento", "document", "documenti"}
    number_tokens = {"n", "nr", "no", "nro", "num", "numero", "number"}

    if compact in doc_exact:
        return "Nr. Doc."

    if any(token in doc_tokens for token in tokens) and any(
        token in number_tokens for token in tokens
    ):
        return "Nr. Doc."

    mc_exact = {
        "mc",
        "m3",
        "metricubi",
        "metrocubo",
        "metrocubi",
        "metri3",
        "metro3",
    }

    if compact in mc_exact:
        return "MC"

    if tokens in (["m", "c"], ["m", "3"]):
        return "MC"

    if "metri" in tokens and "cubi" in tokens:
        return "MC"

    if "metro" in tokens and "cubo" in tokens:
        return "MC"

    kg_terms = {"kg", "kgs", "kilogrammi", "kilogrammo", "chilogrammi", "chilogrammo"}

    if compact in kg_terms:
        return "Kg"

    if any(token in kg_terms for token in tokens):
        return "Kg"

    return None


def scan_header_row(ws: Worksheet, header_row: int) -> dict[str, int]:
    found: dict[str, int] = {}
    found_original_headers: dict[str, str] = {}

    for cell in ws[header_row]:
        canonical_header = identify_required_column(cell.value)

        if canonical_header is None:
            continue

        original_header = normalize_header(cell.value)

        if canonical_header in found:
            previous_header = found_original_headers[canonical_header]
            raise ValueError(
                f"Colonna duplicata o ambigua trovata nella riga {header_row}: "
                f"'{previous_header}' e '{original_header}' sono entrambe riconosciute come "
                f"'{canonical_header}'. Il file deve contenerla una sola volta."
            )

        found[canonical_header] = cell.column
        found_original_headers[canonical_header] = original_header

    return found


def find_required_columns(ws: Worksheet) -> HeaderDetectionResult:
    attempts: list[tuple[int, dict[str, int], list[str]]] = []

    for header_row in (1, 2):
        found = scan_header_row(ws, header_row)
        missing = [col for col in REQUIRED_COLUMNS if col not in found]
        attempts.append((header_row, found, missing))

        if not missing:
            return HeaderDetectionResult(columns=found, header_row=header_row)

    details = []

    for header_row, _found, missing in attempts:
        details.append(f"riga {header_row}: mancanti {', '.join(missing)}")

    raise ValueError(
        "Colonne obbligatorie mancanti. "
        "Il codice cerca le intestazioni solo nella riga 1 o nella riga 2. "
        "Sono accettate anche varianti come 'nr. doc.', 'n. doc.', 'mc', "
        "'metri cubi', 'kg' e 'KG'. "
        + "; ".join(details)
        + "."
    )


def choose_sheet(workbook, sheet_name: str | None) -> Worksheet:
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"Il foglio '{sheet_name}' non esiste nel file. "
                f"Fogli disponibili: {available}."
            )

        return workbook[sheet_name]

    return workbook.worksheets[0]


def parse_doc_number_value(
    value: Any,
    *,
    row: int,
    column_name: str = "Nr. Doc.",
) -> int:
    if is_blank(value):
        raise ValueError(f"Valore vuoto in riga {row}, colonna '{column_name}'.")

    if isinstance(value, bool):
        raise ValueError(
            f"Valore booleano non valido in riga {row}, colonna '{column_name}': {value!r}"
        )

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if value.is_integer():
            return int(value)

        text = format(value, ".15g")
    else:
        text = str(value).strip().replace("\xa0", "").replace(" ", "")

        if text.startswith("="):
            raise ValueError(
                f"Formula non convertibile in riga {row}, colonna '{column_name}': {value!r}. "
                "La colonna deve contenere valori numerici, non formule."
            )

        text = text.replace("'", "")

    negative = False

    if text.startswith("+"):
        text = text[1:]
    elif text.startswith("-"):
        negative = True
        text = text[1:]

    if "e" in text.lower():
        try:
            number = Decimal(text)
        except InvalidOperation as exc:
            raise ValueError(
                f"Impossibile convertire il Nr. Doc. in riga {row}: {value!r}"
            ) from exc

        if number == number.to_integral_value():
            result = int(number)
            return -result if negative else result

        text = format(float(number), ".15f").rstrip("0").rstrip(".")

    if not re.fullmatch(r"\d+(?:[\.,]\d+)*", text):
        raise ValueError(f"Impossibile convertire il Nr. Doc. in riga {row}: {value!r}")

    digits_only = re.sub(r"[\.,]", "", text)
    result = int(digits_only)

    return -result if negative else result


def parse_decimal_value(value: Any, *, row: int, column_name: str) -> Decimal:
    if is_blank(value):
        raise ValueError(f"Valore vuoto in riga {row}, colonna '{column_name}'.")

    if isinstance(value, bool):
        raise ValueError(
            f"Valore booleano non valido in riga {row}, colonna '{column_name}': {value!r}"
        )

    if isinstance(value, int):
        return Decimal(value)

    if isinstance(value, float):
        return Decimal(str(value))

    text = str(value).strip().replace("\xa0", "").replace(" ", "")

    if text.startswith("="):
        raise ValueError(
            f"Formula non convertibile in riga {row}, colonna '{column_name}': {value!r}. "
            "La colonna deve contenere valori numerici, non formule."
        )

    text = text.replace("€", "").replace("'", "")

    negative = False

    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]

    if text.startswith("+"):
        text = text[1:]
    elif text.startswith("-"):
        negative = True
        text = text[1:]

    if not re.fullmatch(r"\d+(?:[\.,]\d+)*", text):
        raise ValueError(
            f"Impossibile convertire il valore numerico in riga {row}, "
            f"colonna '{column_name}': {value!r}"
        )

    comma_positions = [m.start() for m in re.finditer(",", text)]
    dot_positions = [m.start() for m in re.finditer(r"\.", text)]

    if comma_positions and dot_positions:
        last_comma = comma_positions[-1]
        last_dot = dot_positions[-1]
        decimal_pos = max(last_comma, last_dot)

        integer_part = text[:decimal_pos].replace(",", "").replace(".", "")
        decimal_part = text[decimal_pos + 1 :]

        normalized = f"{integer_part}.{decimal_part}"

    elif comma_positions or dot_positions:
        separator = "," if comma_positions else "."
        parts = text.split(separator)

        if len(parts) == 2:
            left, right = parts

            if len(right) <= 2:
                normalized = f"{left}.{right}"
            elif len(right) == 3 and len(left) <= 3:
                normalized = f"{left}{right}"
            else:
                normalized = f"{left}.{right}"

        else:
            last_part = parts[-1]

            if len(last_part) <= 2:
                normalized = "".join(parts[:-1]) + "." + last_part
            elif all(len(part) == 3 for part in parts[1:]):
                normalized = "".join(parts)
            else:
                normalized = "".join(parts[:-1]) + "." + last_part

    else:
        normalized = text

    if negative:
        normalized = "-" + normalized

    try:
        return Decimal(normalized)
    except InvalidOperation as exc:
        raise ValueError(
            f"Impossibile convertire il valore numerico in riga {row}, "
            f"colonna '{column_name}': {value!r}"
        ) from exc


def round_to_2_decimals(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def format_decimal_it(value: Decimal) -> str:
    value = round_to_2_decimals(value)

    sign = "-" if value < 0 else ""
    value = abs(value)

    integer_part = int(value)
    decimal_part = int((value - Decimal(integer_part)) * Decimal("100"))

    integer_text = f"{integer_part:,}".replace(",", ".")
    decimal_text = f"{decimal_part:02d}"

    return f"{sign}{integer_text},{decimal_text}"


def format_integer_it(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def normalize_original_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))

        return format(value, ".15g")

    return str(value).strip()


def build_runs(values: list[int | None]) -> list[tuple[int, int, int | None]]:
    runs: list[tuple[int, int, int | None]] = []
    start = 0

    while start < len(values):
        end = start + 1

        while end < len(values) and values[end] == values[start]:
            end += 1

        runs.append((start, end, values[start]))
        start = end

    return runs


def nearest_previous_non_blank_run_value(
    runs: list[tuple[int, int, int | None]],
    idx: int,
) -> int | None:
    for previous_idx in range(idx - 1, -1, -1):
        value = runs[previous_idx][2]
        if value is not None:
            return value

    return None


def nearest_next_non_blank_run_value(
    runs: list[tuple[int, int, int | None]],
    idx: int,
) -> int | None:
    for next_idx in range(idx + 1, len(runs)):
        value = runs[next_idx][2]
        if value is not None:
            return value

    return None


def trailing_zero_candidates(value: int, max_zeroes: int = 6) -> list[int]:
    if value < 0:
        return []

    text = str(value)
    return [int(text + ("0" * zeroes)) for zeroes in range(1, max_zeroes + 1)]


def correct_truncated_doc_numbers(
    doc_numbers: list[int | None],
    *,
    first_data_row: int,
    max_neighbor_gap: int = 200,
) -> tuple[list[int | None], list[DocCorrection]]:
    corrected = doc_numbers[:]
    corrections: list[DocCorrection] = []
    runs = build_runs(doc_numbers)

    for idx, (start, end, raw_value) in enumerate(runs):
        if raw_value is None:
            continue

        previous_value = nearest_previous_non_blank_run_value(runs, idx)
        next_value = nearest_next_non_blank_run_value(runs, idx)

        candidates: list[tuple[int, Decimal, int, str]] = []

        if previous_value is not None and next_value is not None:
            low = min(previous_value, next_value)
            high = max(previous_value, next_value)

            raw_is_between_neighbors = low < raw_value < high

            if (
                not raw_is_between_neighbors
                and abs(previous_value - next_value) <= max_neighbor_gap
            ):
                expected_midpoint = Decimal(previous_value + next_value) / Decimal(2)

                for candidate in trailing_zero_candidates(raw_value):
                    if low < candidate < high:
                        distance = abs(Decimal(candidate) - expected_midpoint)
                        candidates.append(
                            (
                                0,
                                distance,
                                candidate,
                                "between_previous_and_next",
                            )
                        )

        if previous_value is not None:
            for candidate in trailing_zero_candidates(raw_value):
                if abs(candidate - previous_value) == 1 and abs(raw_value - previous_value) > 50:
                    candidates.append(
                        (
                            1,
                            Decimal(0),
                            candidate,
                            "adjacent_to_previous",
                        )
                    )

        if next_value is not None:
            for candidate in trailing_zero_candidates(raw_value):
                if abs(candidate - next_value) == 1 and abs(raw_value - next_value) > 50:
                    candidates.append(
                        (
                            1,
                            Decimal(0),
                            candidate,
                            "adjacent_to_next",
                        )
                    )

        if not candidates:
            continue

        candidates.sort(key=lambda item: (item[0], item[1], item[2]))
        _, _, corrected_value, rule = candidates[0]

        for pos in range(start, end):
            corrected[pos] = corrected_value

        corrections.append(
            DocCorrection(
                first_excel_row=first_data_row + start,
                last_excel_row=first_data_row + end - 1,
                original_value=raw_value,
                corrected_value=corrected_value,
                previous_distinct_value=previous_value,
                next_distinct_value=next_value,
                rule=rule,
                affected_rows=end - start,
            )
        )

    return corrected, corrections


def sort_data_rows_by_doc(
    ws: Worksheet,
    *,
    nr_doc_col: int,
    first_data_row: int,
) -> None:
    max_row = ws.max_row
    max_col = ws.max_column

    rows: list[tuple[int, int, int, list[Any]]] = []

    for row in range(first_data_row, max_row + 1):
        doc_value = ws.cell(row=row, column=nr_doc_col).value

        if is_blank(doc_value):
            sort_group = 1
            numeric_doc_value = 0
        else:
            sort_group = 0
            numeric_doc_value = int(doc_value)

        row_values = [
            ws.cell(row=row, column=col).value
            for col in range(1, max_col + 1)
        ]

        rows.append((sort_group, numeric_doc_value, row, row_values))

    rows.sort(key=lambda item: (item[0], item[1], item[2]))

    for output_row, (_sort_group, _doc_value, _original_row, row_values) in enumerate(
        rows,
        start=first_data_row,
    ):
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=output_row, column=col).value = value


def clean_excel_file(
    input_path: str | Path,
    output_path: str | Path,
    sheet_name: str | None = None,
) -> CleaningReport:
    input_path = Path(input_path)
    output_path = Path(output_path)

    if not input_path.exists():
        raise FileNotFoundError(f"File non trovato: {input_path}")

    try:
        workbook = load_workbook(input_path)
    except Exception as exc:
        raise ValueError(
            "Non è stato possibile leggere il file Excel. "
            "Verifica che sia un file .xlsx valido e non protetto da password."
        ) from exc

    ws = choose_sheet(workbook, sheet_name)

    header_detection = find_required_columns(ws)
    columns = header_detection.columns
    header_row = header_detection.header_row

    nr_doc_col = columns["Nr. Doc."]
    mc_col = columns["MC"]
    kg_col = columns["Kg"]

    first_data_row = header_row + 1

    if ws.max_row < first_data_row:
        raise ValueError("Il file non contiene righe dati sotto l'intestazione.")

    doc_numbers: list[int | None] = []

    mc_corrections: list[ValueCorrection] = []
    kg_corrections: list[ValueCorrection] = []

    for row in range(first_data_row, ws.max_row + 1):
        doc_cell = ws.cell(row=row, column=nr_doc_col)
        mc_cell = ws.cell(row=row, column=mc_col)
        kg_cell = ws.cell(row=row, column=kg_col)

        original_doc_value = doc_cell.value
        original_mc_value = mc_cell.value
        original_kg_value = kg_cell.value

        if is_blank(original_doc_value):
            doc_number = None
        else:
            doc_number = parse_doc_number_value(
                original_doc_value,
                row=row,
                column_name="Nr. Doc.",
            )

        if not is_blank(original_mc_value):
            mc_value = parse_decimal_value(
                original_mc_value,
                row=row,
                column_name="MC",
            )

            mc_value_rounded = round_to_2_decimals(mc_value)
            mc_clean_text = format_decimal_it(mc_value_rounded)
            original_mc_text = normalize_original_text(original_mc_value)

            if original_mc_text != mc_clean_text:
                mc_corrections.append(
                    ValueCorrection(
                        excel_row=row,
                        column_name="MC",
                        original_value=original_mc_value,
                        corrected_value=mc_clean_text,
                    )
                )

            mc_cell.value = float(mc_value_rounded)

        if not is_blank(original_kg_value):
            kg_value = parse_decimal_value(
                original_kg_value,
                row=row,
                column_name="Kg",
            )

            kg_value_rounded = round_to_2_decimals(kg_value)
            kg_clean_text = format_decimal_it(kg_value_rounded)
            original_kg_text = normalize_original_text(original_kg_value)

            if original_kg_text != kg_clean_text:
                kg_corrections.append(
                    ValueCorrection(
                        excel_row=row,
                        column_name="Kg",
                        original_value=original_kg_value,
                        corrected_value=kg_clean_text,
                    )
                )

            kg_cell.value = float(kg_value_rounded)

        doc_numbers.append(doc_number)

    corrected_doc_numbers, nr_doc_corrections = correct_truncated_doc_numbers(
        doc_numbers,
        first_data_row=first_data_row,
    )

    for output_row, corrected_doc in enumerate(
        corrected_doc_numbers,
        start=first_data_row,
    ):
        ws.cell(row=output_row, column=nr_doc_col).value = corrected_doc

    sort_data_rows_by_doc(
        ws,
        nr_doc_col=nr_doc_col,
        first_data_row=first_data_row,
    )

    for row in range(first_data_row, ws.max_row + 1):
        ws.cell(row=row, column=nr_doc_col).number_format = NR_DOC_NUMBER_FORMAT
        ws.cell(row=row, column=mc_col).number_format = DECIMAL_NUMBER_FORMAT
        ws.cell(row=row, column=kg_col).number_format = DECIMAL_NUMBER_FORMAT

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    data_rows = ws.max_row - first_data_row + 1

    return CleaningReport(
        sheet_name=ws.title,
        header_row=header_row,
        data_rows=data_rows,
        nr_doc_corrections=nr_doc_corrections,
        mc_corrections=mc_corrections,
        kg_corrections=kg_corrections,
    )


def make_default_output_path(input_path: str | Path) -> Path:
    input_path = Path(input_path)

    return input_path.with_name(f"{input_path.stem}_pulito{input_path.suffix}")


def print_report(report: CleaningReport) -> None:
    nr_doc_rows = sum(item.affected_rows for item in report.nr_doc_corrections)

    print("Pulizia completata.")
    print(f"Foglio pulito: {report.sheet_name}")
    print(f"Riga intestazioni: {report.header_row}")
    print(f"Righe dati: {report.data_rows}")
    print(f"Gruppi Nr. Doc. corretti: {len(report.nr_doc_corrections)}")
    print(f"Righe Nr. Doc. corrette: {nr_doc_rows}")
    print(f"Celle MC normalizzate/corrette: {len(report.mc_corrections)}")
    print(f"Celle Kg normalizzate/corrette: {len(report.kg_corrections)}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Pulisce le colonne Nr. Doc., MC e Kg nei file Excel Rifiuti Nave."
    )

    parser.add_argument(
        "input_path",
        help="Percorso del file Excel da pulire.",
    )

    parser.add_argument(
        "output_path",
        nargs="?",
        help="Percorso del file Excel pulito. Se omesso, crea un file con suffisso _pulito.",
    )

    parser.add_argument(
        "--sheet",
        dest="sheet_name",
        default=None,
        help="Nome del foglio da pulire. Se omesso, viene pulito il primo foglio.",
    )

    args = parser.parse_args()

    input_path = Path(args.input_path)
    output_path = Path(args.output_path) if args.output_path else make_default_output_path(input_path)

    report = clean_excel_file(
        input_path=input_path,
        output_path=output_path,
        sheet_name=args.sheet_name,
    )

    print_report(report)
    print(f"File salvato in: {output_path}")


if __name__ == "__main__":
    main()
