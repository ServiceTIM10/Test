from __future__ import annotations

import tempfile
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

from pulisci_file_rifiuti import clean_excel_file


st.set_page_config(
    page_title="Pulizia file rifiuti nave",
    page_icon="🚢",
    layout="centered",
)


def get_sheet_names(uploaded_file) -> list[str]:
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=False)
    sheet_names = workbook.sheetnames
    uploaded_file.seek(0)
    return sheet_names


def format_excel_rows(first_row: int, last_row: int) -> str:
    if first_row == last_row:
        return str(first_row)

    return f"{first_row}-{last_row}"


def render_nr_doc_report(corrections, max_rows: int = 10) -> None:
    st.subheader("Correzioni Nr. Doc.")

    total_groups = len(corrections)
    total_rows = sum(item.affected_rows for item in corrections)

    col1, col2 = st.columns(2)
    col1.metric("Gruppi corretti", total_groups)
    col2.metric("Righe corrette", total_rows)

    if not corrections:
        st.info("Non sono stati rilevati valori troncati nella colonna Nr. Doc.")
        return

    preview_rows = []

    for item in corrections[:max_rows]:
        preview_rows.append(
            {
                "Righe Excel": format_excel_rows(item.first_excel_row, item.last_excel_row),
                "Valore originale": item.original_value,
                "Valore corretto": item.corrected_value,
                "Righe impattate": item.affected_rows,
            }
        )

    st.dataframe(preview_rows, use_container_width=True, hide_index=True)

    if len(corrections) > max_rows:
        st.caption(
            f"Mostrate le prime {max_rows} correzioni su {len(corrections)} gruppi corretti."
        )


def render_value_report(column_name: str, corrections, max_rows: int = 10) -> None:
    st.subheader(f"Correzioni {column_name}")

    st.metric("Celle corrette / normalizzate", len(corrections))

    if not corrections:
        st.info(f"Non sono state rilevate correzioni nella colonna {column_name}.")
        return

    preview_rows = []

    for item in corrections[:max_rows]:
        preview_rows.append(
            {
                "Riga Excel": item.excel_row,
                "Valore originale": item.original_value,
                "Valore corretto": item.corrected_value,
            }
        )

    st.dataframe(preview_rows, use_container_width=True, hide_index=True)

    if len(corrections) > max_rows:
        st.caption(
            f"Mostrate le prime {max_rows} correzioni su {len(corrections)} celle corrette."
        )


st.title("Pulizia file Excel - Rifiuti Nave 🚢")

st.markdown(
    """
Questa applicazione consente di caricare un file Excel relativo ai **Rifiuti Nave**, correggere automaticamente le colonne **Nr. Doc.**, **MC** e **Kg**, ordinare il database e scaricare il file pulito.

Il file viene elaborato temporaneamente durante la sessione e non viene salvato in modo permanente dall'applicazione.
"""
)

with st.expander("Cosa fa l'app", expanded=False):
    st.markdown(
        """
L'app esegue queste operazioni:

1. cerca le intestazioni **Nr. Doc.**, **MC** e **Kg** nella **riga 1**;
2. se non le trova tutte nella riga 1, le cerca nella **riga 2**;
3. pulisce solo le colonne **Nr. Doc.**, **MC** e **Kg**;
4. converte **Nr. Doc.** in numero intero;
5. gestisce anche i valori di **Nr. Doc.** in formato inglese, ad esempio `5,645` oppure `5.645`;
6. corregge i casi troncati quando coerenti con i valori vicini, ad esempio:
   - `1` → `1.000`;
   - `161` → `1.610`;
   - `25` → `2.500`;
   - `5.64` → `5.640`;
7. converte **MC** e **Kg** in valori numerici con due decimali;
8. applica il formato numerico Excel;
9. ordina l'intera tabella in ordine crescente sulla base di **Nr. Doc.**;
10. restituisce un nuovo file Excel pulito.
"""
    )

with st.expander("Condizioni di funzionamento ed errori", expanded=False):
    st.markdown(
        """
L'app funziona se:

- il file caricato è in formato **.xlsx**;
- le intestazioni sono nella **riga 1** oppure nella **riga 2**;
- il foglio contiene le colonne **Nr. Doc.**, **MC** e **Kg**;
- i valori presenti in queste tre colonne sono numerici oppure testi numerici convertibili;
- sotto la riga delle intestazioni è presente almeno una riga dati.

L'app restituisce errore se:

- il file non è un `.xlsx` valido;
- il file è protetto o non leggibile;
- il nome del foglio indicato non esiste;
- una o più colonne obbligatorie non sono presenti né in riga 1 né in riga 2;
- una delle colonne obbligatorie è duplicata;
- una cella di **Nr. Doc.**, **MC** o **Kg** è vuota;
- una cella contiene una formula invece di un valore numerico;
- una cella contiene un valore non convertibile in numero.

Le altre colonne non vengono pulite né trasformate. Tuttavia vengono riordinate insieme alla relativa riga, perché l'intero foglio viene ordinato in base a **Nr. Doc.**.
"""
    )

uploaded_file = st.file_uploader(
    "Carica il file Excel da pulire",
    type=["xlsx"],
    accept_multiple_files=False,
)

if uploaded_file is None:
    st.warning("Carica un file Excel per iniziare.")
    st.stop()

st.info(f"File caricato: {uploaded_file.name}")

try:
    sheet_names = get_sheet_names(uploaded_file)
except Exception as exc:
    st.error(
        "Non è stato possibile leggere il file Excel. "
        "Verifica che sia un file .xlsx valido e non protetto."
    )
    with st.expander("Dettaglio tecnico errore"):
        st.exception(exc)
    st.stop()

st.success(f"Fogli disponibili: {', '.join(sheet_names)}")

sheet_name = st.text_input(
    "Nome del foglio da pulire",
    value="",
    placeholder="Lascia vuoto per usare il primo foglio del file",
    help="Se lasci il campo vuoto, verrà pulito il primo foglio del file Excel.",
)

if sheet_name.strip():
    st.caption(f"Foglio selezionato manualmente: **{sheet_name.strip()}**")
else:
    st.caption(f"Campo vuoto: verrà pulito il primo foglio, cioè **{sheet_names[0]}**.")

if st.button("Pulisci file", type="primary"):
    with st.spinner("Pulizia del file in corso..."):
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                tmpdir_path = Path(tmpdir)

                safe_input_filename = Path(uploaded_file.name).name
                input_path = tmpdir_path / safe_input_filename
                output_filename = f"pulito_{safe_input_filename}"
                output_path = tmpdir_path / output_filename

                uploaded_file.seek(0)
                input_path.write_bytes(uploaded_file.getbuffer())

                report = clean_excel_file(
                    input_path=input_path,
                    output_path=output_path,
                    sheet_name=sheet_name.strip() or None,
                )

                output_bytes = output_path.read_bytes()

            st.success("File pulito correttamente.")

            col1, col2, col3 = st.columns(3)
            col1.metric("Foglio pulito", report.sheet_name)
            col2.metric("Riga intestazioni", report.header_row)
            col3.metric("Righe dati", report.data_rows)

            with st.expander("Report elaborazione", expanded=True):
                render_nr_doc_report(report.nr_doc_corrections, max_rows=10)
                st.divider()
                render_value_report("MC", report.mc_corrections, max_rows=10)
                st.divider()
                render_value_report("Kg", report.kg_corrections, max_rows=10)

            st.download_button(
                label="Scarica file pulito",
                data=output_bytes,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except KeyError:
            st.error(
                "Il foglio indicato non esiste nel file. "
                "Controlla il nome del foglio e riprova."
            )

        except ValueError as exc:
            st.error(str(exc))

        except FileNotFoundError as exc:
            st.error(str(exc))

        except Exception as exc:
            st.error("Si è verificato un errore inatteso durante la pulizia del file.")
            with st.expander("Dettaglio tecnico errore"):
                st.exception(exc)
